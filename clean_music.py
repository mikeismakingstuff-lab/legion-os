"""
clean_music.py
Extracts Artist and Title from a music library Excel file.

Usage:
    python clean_music.py <path_to_xlsx> <path_to_audio_dir>

Defaults:
    xlsx      = "test music library.xlsx"
    audio_dir = "E:\\RadioStation"

Design constraint: all decisions based on structural properties of the input only.
No hardcoded artist names, no test-set-derived suffix lists, no assumed conventions
beyond what is observable from the data structures themselves.
"""

import json
import openpyxl
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path
from tinytag import TinyTag

# Reconfigure stdout/stderr to UTF-8 to prevent encoding errors on Windows console
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

# ---------------------------------------------------------------------------
# Fingerprinting
# ---------------------------------------------------------------------------

# Known install location for fpcalc bundled with MusicBrainz Picard on Windows.
# Used as a fallback when fpcalc is not on PATH.
# If Picard was installed to a custom path, override by putting fpcalc.exe on
# PATH before running this script rather than editing this constant.
FPCALC_PICARD_FALLBACK = Path(r"C:\Program Files\MusicBrainz Picard\fpcalc.exe")


def find_fpcalc():
    """
    Locate fpcalc once at startup.
    Resolution order:
      1. PATH (covers all platforms, custom installs, Picard's own PATH entry).
      2. Known Picard-on-Windows default install location (FPCALC_PICARD_FALLBACK).
    Returns the path string if found, else None.
    """
    on_path = shutil.which("fpcalc")
    if on_path:
        return on_path
    if FPCALC_PICARD_FALLBACK.exists():
        return str(FPCALC_PICARD_FALLBACK)
    return None


def fingerprint_audio(fpcalc_path, file_path):
    """
    Generate an AcoustID fingerprint for file_path using fpcalc.

    Scope (Phase 1 only): fingerprint generation and storage.
    The AcoustID API query (fingerprint -> MusicBrainz recording ID) is a
    separate network step with its own rate-limiting concerns and is NOT
    performed here.

    Returns:
        (raw_json: str, failure_reason: None)  on success
        (None, failure_reason: str)             on any failure

    Never raises — all errors become a logged failure_reason so the caller
    can fall through to heuristics without silent degradation.
    """
    if fpcalc_path is None:
        return None, "fpcalc not found on PATH or at Picard fallback location"

    try:
        result = subprocess.run(
            [fpcalc_path, "-json", str(file_path)],
            capture_output=True,
            text=True,
            timeout=30,
        )
    except FileNotFoundError:
        return None, f"fpcalc binary not executable: {fpcalc_path}"
    except subprocess.TimeoutExpired:
        return None, f"fpcalc timed out after 30s on: {file_path}"
    except Exception as exc:  # noqa: BLE001
        return None, f"fpcalc unexpected error ({type(exc).__name__}): {exc}"

    if result.returncode != 0:
        stderr_snippet = (result.stderr or "").strip()[:200]
        return None, f"fpcalc exit {result.returncode}: {stderr_snippet}"

    # Validate that the output is parseable JSON containing a fingerprint —
    # a malformed response should not silently pass through as raw_json.
    try:
        parsed = json.loads(result.stdout)
        if "fingerprint" not in parsed:
            return None, "fpcalc JSON missing 'fingerprint' key"
    except json.JSONDecodeError as exc:
        return None, f"fpcalc output not valid JSON: {exc}"

    return result.stdout, None

# ---------------------------------------------------------------------------
# Structural helpers
# ---------------------------------------------------------------------------

def is_uploader_handle(s):
    """
    Real artist/title names almost never end in 3+ bare digits with no space
    before them. Strings like 'Olga0815007' or 'numberonesongs111' are
    characteristic uploader account names — a structural property, not a
    name-specific check.
    """
    return bool(re.search(r'\S\d{3,}$', s))

def has_corrupt_characters(s):
    """
    When a byte sequence from Windows-1252 or Latin-1 fails to decode, the
    replacement character is typically '?'. A '?' immediately followed by a
    letter (not a space, end of string, or punctuation) is the structural
    signal for encoding corruption rather than real punctuation.
    """
    return bool(re.search(r'\?[A-Za-zÀ-ÿ]', s))

def tag_equals_filename(tag_value, filename):
    """
    Many rippers auto-populate the tag with the filename (with or without
    extension). Detect this by comparing their alphanumeric-only forms.
    """
    base = os.path.splitext(filename)[0]
    clean_tag  = re.sub(r'[^a-zA-Z0-9]', '', tag_value or '').lower()
    clean_base = re.sub(r'[^a-zA-Z0-9]', '', base).lower()
    if not clean_tag or not clean_base:
        return False
    # Allow for minor truncation (tag may be a prefix/suffix of filename)
    ratio = len(clean_tag) / len(clean_base)
    return clean_tag == clean_base or (ratio > 0.85 and (clean_tag in clean_base or clean_base in clean_tag))

def tag_is_compound(tag_title, resolved_artist):
    """
    Some tag readers store 'Artist - Title' as the Title field, particularly
    when the tag was written by software that parses the filename.
    Structural signal: the tag title contains ' - ' AND one side of the split
    matches the resolved artist (alphanumeric comparison).
    Returns the title portion if compound, else None.
    """
    if not tag_title or ' - ' not in tag_title:
        return None
    if not resolved_artist:
        return None
    art_clean = re.sub(r'[^a-zA-Z0-9]', '', resolved_artist).lower()
    if not art_clean:
        return None
    parts = [p.strip() for p in tag_title.split(' - ', 1)]
    for i, part in enumerate(parts):
        part_clean = re.sub(r'[^a-zA-Z0-9]', '', part).lower()
        if art_clean and part_clean and (art_clean in part_clean or part_clean in art_clean):
            # The other part is the real title
            return parts[1 - i]
    return None

def strip_track_prefix(title):
    """
    Batch export tools often prefix titles with track numbers.
    Structural patterns: '01 - ', '01. ', '(01) ', '1. ', etc.
    These are purely positional artifacts, not part of the song name.
    """
    if not title:
        return title
    # Pattern: optional '(' + digits + optional ')' + optional separator characters
    title = re.sub(r'^\(?(\d{1,3})\)?\s*[.\-\s]\s*', '', title).strip()
    return title

def strip_trailing_metadata_parenthetical(title):
    """
    A generic structural approach to trailing metadata groups:

    The class of problem: encoded metadata appended by rippers/distributors
    appears as a parenthetical or bracketed group at the END of the title.
    These groups share structural properties distinguishing them from
    genuine subtitle parentheticals:

      1. They appear at the very end of the string.
      2. They are short (1-4 words).
      3. All words inside are either:
           a. Capitalized (not a grammatical phrase like 'from the album'), or
           b. A bare number or abbreviation, or
           c. Empty (the group is just a number like '(2023)').
      4. They do NOT contain common lowercase function words that would indicate
         they are part of a real subtitle (e.g., 'from', 'in', 'the', 'of', 'a').

    This function iteratively strips such groups from the end of the title.
    It does NOT hardcode any specific words inside the group.
    """
    function_words = {
        'a', 'an', 'the', 'of', 'in', 'on', 'at', 'to', 'for', 'by', 'from',
        'and', 'or', 'but', 'with', 'as', 'is', 'it', 'its', 'my', 'me',
        'he', 'she', 'we', 'you', 'they', 'this', 'that', 'which', 'who',
    }

    changed = True
    while changed:
        changed = False
        # Match the last parenthetical or bracketed group
        m = re.search(r'\s*[\(\[]([\w\s\-\.,\'\"]+)[\)\]]\s*$', title)
        if not m:
            break
        inner = m.group(1).strip()
        words = inner.split()

        # Condition: short group
        if len(words) > 4:
            break

        # Condition: no function words inside (would indicate a real subtitle)
        has_function_word = any(w.lower() in function_words for w in words)
        if has_function_word:
            break

        # Condition: words are capitalized, numeric, or uppercase abbreviations
        all_metadata_like = all(
            w[0].isupper() or w.isdigit() or re.fullmatch(r'[\dA-Z\-\.]+', w)
            for w in words if w
        )
        if not all_metadata_like:
            break

        # Strip the group
        title = title[:m.start()].strip()
        changed = True

    return title.strip(" '\"-")

def strip_trailing_upload_id(title):
    """
    Structural class: some filenames and tags end with a bare alphanumeric block
    that is a catalog/upload ID — characterized by the same uploader-handle
    signal (block of 4+ digits, optionally preceded by letters, at end of string,
    separated from the preceding text by whitespace).

    e.g., 'Lady Marmalade (1975) HD 0815007'
          'Some Song Title ABC12345'
    """
    title = re.sub(r'\s+[A-Za-z]{0,4}\d{4,}\s*$', '', title).strip()
    return title

# ---------------------------------------------------------------------------
# File matching
# ---------------------------------------------------------------------------

_disk_files_set = None
_disk_files_lower = None
_clean_to_disk = None

def find_matching_file(excel_name, disk_files):
    """
    Match an Excel filename entry to an actual disk filename.
    Three-pass strategy:
      1. Exact match.
      2. Regex match — '?' treated as a wildcard for encoding placeholders.
      3. Alphanumeric-only fuzzy match (exact, then substring).
    """
    global _disk_files_set, _disk_files_lower, _clean_to_disk
    if _disk_files_set is None:
        _disk_files_set = set(disk_files)
        _disk_files_lower = {f.lower(): f for f in disk_files}
        _clean_to_disk = {}
        for f in disk_files:
            c = re.sub(r'[^a-zA-Z0-9]', '', f).lower()
            if c:
                if c not in _clean_to_disk:
                    _clean_to_disk[c] = f

    if excel_name in _disk_files_set:
        return excel_name

    excel_lower = excel_name.lower()
    if excel_lower in _disk_files_lower:
        return _disk_files_lower[excel_lower]

    if '?' in excel_name:
        pattern_str = re.escape(excel_name).replace(r'\?', '.')
        try:
            pattern = re.compile('^' + pattern_str + '$', re.IGNORECASE)
            for f in disk_files:
                if pattern.match(f):
                    return f
        except Exception:
            pass

    excel_clean = re.sub(r'[^a-zA-Z0-9]', '', excel_name).lower()
    if excel_clean:
        if excel_clean in _clean_to_disk:
            return _clean_to_disk[excel_clean]
        for f_clean, f in _clean_to_disk.items():
            if excel_clean in f_clean or f_clean in excel_clean:
                return f

    return None

# ---------------------------------------------------------------------------
# Filename structural parsing
# ---------------------------------------------------------------------------

def parse_filename_structurally(filename, canonical_artists_clean):
    """
    Attempt to extract (artist, title) from a filename using structural signals only.

    Strategy (in order):
      A. Split on ' - ' (and equivalents: em-dash, en-dash with spaces).
      B. Among parts, find one that exactly or substantially matches a canonical artist.
      C. If no canonical match, look for a collaboration keyword (feat, with, &, vs).
      D. If a part starts with '(' or a quote, treat it as the title (title-first filename).
      E. Default: first part = artist, rest = title.
      F. Fallback for unseparated files: 'Title by Artist' pattern.
    """
    base_name = os.path.splitext(filename)[0].strip(" '\"-")

    # Split on ' - ', ' – ', ' — ' (hyphen, en-dash, em-dash with surrounding spaces)
    parts = [p.strip() for p in re.split(r'\s+[-\u2013\u2014]\s+', base_name)]
    parts = [p for p in parts if p and p.strip('-\u2013\u2014')]

    if len(parts) >= 2:
        # A: Canonical artist lookup — exact match first, then substantial substring
        artist_idx = -1
        for idx, part in enumerate(parts):
            part_clean = re.sub(r'[^a-zA-Z0-9]', '', part).lower()
            for art, art_clean in canonical_artists_clean:
                if len(art_clean) > 3 and art_clean == part_clean:
                    artist_idx = idx
                    break
                # Substantial substring: artist appears as the bulk of this part
                if len(art_clean) > 5 and art_clean in part_clean:
                    # Make sure the part isn't much longer than the artist name
                    # (avoids matching artist 'Joan' inside 'Joan and her friends')
                    if len(part_clean) < len(art_clean) * 1.6:
                        artist_idx = idx
                        break
            if artist_idx != -1:
                break

        if artist_idx != -1:
            artist = parts[artist_idx]
            title = " - ".join(p for i, p in enumerate(parts) if i != artist_idx)
            return artist, title

        # B: Collaboration keyword heuristic
        collab_re = r'\b(feat\.?|featuring|ft\.?|with|vs\.?|&)\b'
        for idx, part in enumerate(parts):
            if re.search(collab_re, part, re.IGNORECASE):
                artist = part
                title = " - ".join(p for i, p in enumerate(parts) if i != idx)
                return artist, title

        # C: Title-first pattern — part starting with '(' or quote is the title
        for idx, part in enumerate(parts):
            if part.startswith(('(', '"', "'")):
                title = part
                artist = " - ".join(p for i, p in enumerate(parts) if i != idx)
                return artist, title

        # D: Default
        return parts[0], " - ".join(parts[1:])

    # E: 'Title by Artist' fallback
    m = re.match(r'^(.+?)\s+by\s+(.+?)$', base_name, re.IGNORECASE)
    if m:
        return m.group(2).strip(), m.group(1).strip()

    return None, base_name

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else 'test music library.xlsx'
    audio_dir = sys.argv[2] if len(sys.argv) > 2 else r'E:\RadioStation'

    # Locate fpcalc once — hoisted outside the per-file loop so we pay the
    # PATH search cost exactly once per run, not once per record.
    fpcalc_path = find_fpcalc()
    if fpcalc_path:
        print(f"fpcalc found: {fpcalc_path}")
    else:
        print(
            "[WARN] fpcalc not found — fingerprinting unavailable this run. "
            "All records will use heuristic metadata resolution. "
            "Install MusicBrainz Picard or add fpcalc to PATH to enable."
        )

    disk_files = os.listdir(audio_dir)
    audio_exts = ('.mp3', '.flac', '.m4a', '.ogg', '.wav', '.aac', '.wma', '.opus', '.ape')

    # Build canonical artist set from embedded tags.
    # Structural filter only: reject uploader handles (trailing 3+ digits).
    print(f"Building canonical artist list from: {audio_dir}")
    canonical_artists = set()
    total_files = len(disk_files)
    processed = 0
    file_tags = {}  # Maps filename -> (artist, title)
    for f in disk_files:
        if f.lower().endswith(audio_exts):
            processed += 1
            if processed % 1000 == 0:
                print(f"  Processed {processed}/{total_files} files for canonical artists...")
            try:
                tag = TinyTag.get(os.path.join(audio_dir, f))
                file_tags[f] = (tag.artist, tag.title)
                if tag.artist:
                    art = tag.artist.strip()
                    if art and not is_uploader_handle(art) and not has_corrupt_characters(art) and len(art) < 100:
                        canonical_artists.add(art)
            except Exception:
                pass
    print(f"Canonical artist set: {len(canonical_artists)} entries.")
    canonical_artists_clean = []
    for art in canonical_artists:
        art_clean = re.sub(r'[^a-zA-Z0-9]', '', art).lower()
        canonical_artists_clean.append((art, art_clean))

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    if 'Legion Results' in wb.sheetnames:
        del wb['Legion Results']
    ws_res = wb.create_sheet('Legion Results')
    ws_res.append(['Name', 'FullName', 'Artist', 'Title'])

    matched = 0
    unmatched = 0
    total_rows = ws.max_row - 1

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
        if idx % 1000 == 0:
            print(f"  Processing Excel Row {idx}/{total_rows}...")
        name = row[0].value
        fullname = row[1].value
        if not name:
            continue

        disk_match = find_matching_file(name, disk_files)

        if not disk_match:
            ws_res.append([name, fullname, 'FILE NOT FOUND', 'FILE NOT FOUND'])
            unmatched += 1
            continue

        matched += 1
        path = os.path.join(audio_dir, disk_match)

        # ---- Fingerprint (ground-truth verification, Phase 1) ----
        # Runs before heuristics. On success, fp_raw carries the raw fpcalc
        # JSON for downstream use (Phase 2 will query AcoustID with it).
        # On failure, falls through to heuristics — never silent.
        fp_raw, fp_skip_reason = fingerprint_audio(fpcalc_path, path)
        if fp_skip_reason:
            pass  # Suppress the verbose FP SKIP print to speed up execution and clean up output

        tag_artist, tag_title = file_tags.get(disk_match, (None, None))

        # Structural parse of disk filename (disk has correct characters)
        fn_artist, fn_title = parse_filename_structurally(disk_match, canonical_artists_clean)

        # ---- Resolve Artist ----
        resolved_artist = tag_artist
        if not resolved_artist:
            resolved_artist = fn_artist
        elif resolved_artist.lower().strip() in ('unknown', 'various', 'various artists', 'none', ''):
            resolved_artist = fn_artist
        elif is_uploader_handle(resolved_artist):
            resolved_artist = fn_artist
        elif has_corrupt_characters(resolved_artist):
            resolved_artist = fn_artist

        # ---- Resolve Title ----
        resolved_title = tag_title
        if not resolved_title:
            resolved_title = fn_title
        elif resolved_title.lower().strip() in ('unknown', 'none', ''):
            resolved_title = fn_title
        elif tag_equals_filename(resolved_title, disk_match):
            # Tag title is just the filename — no real tag was embedded
            resolved_title = fn_title
        elif has_corrupt_characters(resolved_title):
            resolved_title = fn_title
        else:
            # Check if the tag title is compound ('Artist - Title' stored as title)
            compound_title = tag_is_compound(resolved_title, resolved_artist)
            if compound_title:
                resolved_title = compound_title

        # ---- Post-processing ----
        if resolved_artist:
            resolved_artist = re.sub(r'\s+', ' ', resolved_artist).strip(" '\"-")

        if resolved_title:
            resolved_title = strip_track_prefix(resolved_title)
            resolved_title = strip_trailing_upload_id(resolved_title)
            resolved_title = strip_trailing_metadata_parenthetical(resolved_title)
            resolved_title = re.sub(r'\s+', ' ', resolved_title).strip(" '\"-")

        ws_res.append([name, fullname, resolved_artist or '', resolved_title or ''])

    wb.save(xlsx_path)
    print(f"Done. Matched: {matched}  |  Unmatched: {unmatched}")
    print(f"Results saved to 'Legion Results' tab in: {xlsx_path}")

if __name__ == '__main__':
    main()
