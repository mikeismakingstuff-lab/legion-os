"""
rename_music.py
Legion OS — Music Library Renaming & Standardization Utility

This script reads the resolved metadata (Artist and Title) from the Excel library,
safely renames the corresponding audio files on disk to a standard convention
(e.g., 'Artist - Title.ext'), and updates their internal metadata tags.

It features a mandatory dry-run mode, conflict detection, filename sanitization,
transaction logging, and full rollback capability.

Usage:
    python rename_music.py <path_to_xlsx> <path_to_audio_dir> [options]

Options:
    --commit            Execute the renaming and tag updates (default: dry-run only).
    --rollback <log>    Revert a previous rename operation using the transaction log.
    --format <pattern>  Naming format pattern (default: "{artist} - {title}").
"""

import argparse
import datetime
import json
import os
import re
import sys
from pathlib import Path
import openpyxl
import mutagen

# Reconfigure stdout/stderr to UTF-8 to prevent encoding errors on Windows console
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

# Import heuristics and helpers from clean_music if needed
try:
    from clean_music import (
        find_matching_file,
        parse_filename_structurally,
        is_uploader_handle,
        has_corrupt_characters,
        tag_equals_filename,
        tag_is_compound,
        strip_track_prefix,
        strip_trailing_metadata_parenthetical,
        strip_trailing_upload_id,
        find_fpcalc,
    )
except ImportError:
    # Fallback definitions if clean_music.py is not in the same directory
    def find_matching_file(excel_name, disk_files):
        if excel_name in disk_files:
            return excel_name
        excel_clean = re.sub(r'[^a-zA-Z0-9]', '', excel_name).lower()
        if excel_clean:
            for f in disk_files:
                f_clean = re.sub(r'[^a-zA-Z0-9]', '', f).lower()
                if excel_clean == f_clean:
                    return f
            for f in disk_files:
                f_clean = re.sub(r'[^a-zA-Z0-9]', '', f).lower()
                if f_clean and (excel_clean in f_clean or f_clean in excel_clean):
                    return f
        return None

    def parse_filename_structurally(filename, canonical_artists):
        base_name = os.path.splitext(filename)[0].strip(" '\"-")
        parts = [p.strip() for p in re.split(r'\s+[-\u2013\u2014]\s+', base_name)]
        parts = [p for p in parts if p and p.strip('-\u2013\u2014')]
        if len(parts) >= 2:
            return parts[0], " - ".join(parts[1:])
        return None, base_name


def sanitize_filename(filename):
    """
    Sanitize filename to remove Windows-invalid characters: < > : " / \\ | ? *
    and replace them with safe alternatives.
    """
    replacements = {
        '<': '',
        '>': '',
        ':': ' -',
        '"': '',
        '/': '-',
        '\\': '-',
        '|': '-',
        '?': '',
        '*': ''
    }
    for char, repl in replacements.items():
        filename = filename.replace(char, repl)
    # Clean up multiple spaces and strip leading/trailing spaces/dots
    filename = re.sub(r'\s+', ' ', filename).strip()
    filename = filename.rstrip('. ')
    return filename


def update_tags(file_path, artist, title):
    """
    Update the audio file's internal metadata tags using mutagen.
    Supports MP3, FLAC, M4A, Ogg, etc.
    """
    file_path_str = str(file_path)
    try:
        audio = mutagen.File(file_path_str, easy=True)
        if audio is None:
            # If it's an MP3 file, try to initialize ID3 tags
            if file_path_str.lower().endswith('.mp3'):
                from mutagen.mp3 import MP3
                from mutagen.easyid3 import EasyID3
                try:
                    audio_mp3 = MP3(file_path_str)
                    if audio_mp3.tags is None:
                        audio_mp3.add_tags()
                        audio_mp3.save()
                    audio = EasyID3(file_path_str)
                except Exception as e:
                    return False, f"Failed to initialize MP3 tags: {e}"
            else:
                return False, "Unsupported format or unable to parse tags"

        audio['artist'] = artist
        audio['title'] = title
        audio.save()
        return True, None
    except Exception as e:
        return False, str(e)


def execute_rollback(log_path):
    """
    Revert a previous renaming operation using the transaction log.
    """
    log_path = Path(log_path)
    if not log_path.exists():
        print(f"[ERROR] Transaction log not found: {log_path}")
        return False

    try:
        with open(log_path, 'r', encoding='utf-8') as f:
            transaction = json.load(f)
    except Exception as e:
        print(f"[ERROR] Failed to read transaction log: {e}")
        return False

    actions = transaction.get("actions", [])
    if not actions:
        print("[INFO] No actions found in transaction log.")
        return True

    print(f"Starting rollback of {len(actions)} actions from log: {log_path.name}")
    success_count = 0
    fail_count = 0

    # Rollback in reverse order to handle any sequential dependencies correctly
    for action in reversed(actions):
        old_path = Path(action["old_path"])
        new_path = Path(action["new_path"])

        if not new_path.exists():
            print(f"  [SKIP] Current file not found (already moved or deleted?): {new_path}")
            fail_count += 1
            continue

        if old_path.exists():
            print(f"  [CONFLICT] Original path already exists: {old_path}")
            fail_count += 1
            continue

        try:
            # Ensure parent directory exists
            old_path.parent.mkdir(parents=True, exist_ok=True)
            os.rename(new_path, old_path)
            print(f"  [OK] Rolled back: {new_path.name} -> {old_path.name}")
            success_count += 1
        except Exception as e:
            print(f"  [ERROR] Failed to roll back {new_path.name}: {e}")
            fail_count += 1

    print(f"\nRollback complete. Success: {success_count} | Failed: {fail_count}")
    return fail_count == 0


def main():
    parser = argparse.ArgumentParser(description="Rename music files based on cleaned metadata.")
    parser.add_argument("xlsx_path", nargs="?", default="test music library.xlsx", help="Path to Excel library file")
    parser.add_argument("audio_dir", nargs="?", default=r"E:\RadioStation", help="Path to audio directory")
    parser.add_argument("--commit", action="store_true", help="Commit the changes to disk (default is dry-run)")
    parser.add_argument("--rollback", help="Path to a transaction log JSON file to revert changes")
    parser.add_argument("--format", default="{artist} - {title}", help="Naming format pattern (default: '{artist} - {title}')")

    args = parser.parse_args()

    # Handle rollback mode
    if args.rollback:
        success = execute_rollback(args.rollback)
        sys.exit(0 if success else 1)

    xlsx_path = Path(args.xlsx_path)
    audio_dir = Path(args.audio_dir)

    if not xlsx_path.exists():
        print(f"[ERROR] Excel file not found: {xlsx_path}")
        sys.exit(1)
    if not audio_dir.exists() or not audio_dir.is_dir():
        print(f"[ERROR] Audio directory not found: {audio_dir}")
        sys.exit(1)

    print(f"Loading workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)
    
    # Check if 'Legion Results' sheet exists
    if 'Legion Results' not in wb.sheetnames:
        print("[WARN] 'Legion Results' sheet not found in workbook.")
        print("Please run clean_music.py first to generate the cleaned metadata sheet,")
        print("or ensure the workbook has a sheet named 'Legion Results' with 'Name', 'Artist', and 'Title' columns.")
        sys.exit(1)

    ws = wb['Legion Results']
    
    # Find column indices
    headers = [cell.value for cell in ws[1]]
    try:
        name_idx = headers.index('Name')
        artist_idx = headers.index('Artist')
        title_idx = headers.index('Title')
    except ValueError as e:
        print(f"[ERROR] Missing required column in 'Legion Results' sheet: {e}")
        sys.exit(1)

    disk_files = os.listdir(audio_dir)
    audio_exts = ('.mp3', '.flac', '.m4a', '.ogg', '.wav', '.aac', '.wma', '.opus', '.ape')

    proposed_actions = []
    dest_paths = {}  # Maps proposed destination path -> source path (for conflict detection)

    print(f"Scanning rows and matching files in: {audio_dir}")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        excel_name = row[name_idx].value
        artist = row[artist_idx].value
        title = row[title_idx].value

        if not excel_name:
            continue

        # Find matching file on disk
        disk_match = find_matching_file(excel_name, disk_files)
        if not disk_match:
            proposed_actions.append({
                "row": row_idx,
                "excel_name": excel_name,
                "status": "FILE NOT FOUND",
                "reason": "Could not match Excel entry to any file on disk"
            })
            continue

        source_path = audio_dir / disk_match
        ext = source_path.suffix

        # Validate metadata
        if not artist or not title or artist == 'FILE NOT FOUND' or title == 'FILE NOT FOUND':
            proposed_actions.append({
                "row": row_idx,
                "excel_name": excel_name,
                "source_path": str(source_path),
                "status": "MISSING METADATA",
                "reason": f"Missing artist or title (Artist: {artist}, Title: {title})"
            })
            continue

        # Format new filename
        try:
            new_name_base = args.format.format(artist=artist, title=title)
        except KeyError as e:
            print(f"[ERROR] Invalid format pattern. Missing key: {e}")
            sys.exit(1)

        new_name_base = sanitize_filename(new_name_base)
        new_name = f"{new_name_base}{ext}"
        dest_path = audio_dir / new_name

        # Check if no change is needed
        if disk_match == new_name:
            proposed_actions.append({
                "row": row_idx,
                "excel_name": excel_name,
                "source_path": str(source_path),
                "dest_path": str(dest_path),
                "artist": artist,
                "title": title,
                "status": "NO CHANGE",
                "reason": "Filename already matches standard convention"
            })
            continue

        # Handle duplicate destination paths (conflicts)
        if dest_path in dest_paths:
            # Resolve conflict by appending a counter
            counter = 1
            while True:
                resolved_name = f"{new_name_base} ({counter}){ext}"
                resolved_dest_path = audio_dir / resolved_name
                if resolved_dest_path not in dest_paths and not resolved_dest_path.exists():
                    new_name = resolved_name
                    dest_path = resolved_dest_path
                    break
                counter += 1
            
            proposed_actions.append({
                "row": row_idx,
                "excel_name": excel_name,
                "source_path": str(source_path),
                "dest_path": str(dest_path),
                "artist": artist,
                "title": title,
                "status": "CONFLICT RESOLVED",
                "reason": f"Destination conflict resolved by appending suffix: {new_name}"
            })
        else:
            proposed_actions.append({
                "row": row_idx,
                "excel_name": excel_name,
                "source_path": str(source_path),
                "dest_path": str(dest_path),
                "artist": artist,
                "title": title,
                "status": "PENDING",
                "reason": "Ready to rename"
            })

        dest_paths[dest_path] = source_path

    # Print summary
    total = len(proposed_actions)
    pending = sum(1 for a in proposed_actions if a["status"] in ("PENDING", "CONFLICT RESOLVED"))
    no_change = sum(1 for a in proposed_actions if a["status"] == "NO CHANGE")
    missing = sum(1 for a in proposed_actions if a["status"] == "MISSING METADATA")
    not_found = sum(1 for a in proposed_actions if a["status"] == "FILE NOT FOUND")

    print("\n" + "="*50)
    print("RENAME SUMMARY")
    print("="*50)
    print(f"Total entries processed: {total}")
    print(f"Ready to rename:        {pending}")
    print(f"No change required:     {no_change}")
    print(f"Missing metadata:       {missing}")
    print(f"Files not found:        {not_found}")
    print("="*50 + "\n")

    if pending == 0:
        print("[INFO] No files require renaming.")
        sys.exit(0)

    # Dry-run mode
    if not args.commit:
        print("[DRY-RUN] Showing first 20 proposed changes:")
        count = 0
        for action in proposed_actions:
            if action["status"] in ("PENDING", "CONFLICT RESOLVED"):
                old_name = Path(action["source_path"]).name
                new_name = Path(action["dest_path"]).name
                print(f"  Row {action['row']}: {old_name} -> {new_name} ({action['status']})")
                count += 1
                if count >= 20:
                    break
        print(f"\n[DRY-RUN] To execute these {pending} renames and update tags, run with --commit.")
        sys.exit(0)

    # Commit mode
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    transaction_log_path = Path(f"rename_transaction_{timestamp}.json")
    
    transaction_log = {
        "timestamp": timestamp,
        "xlsx_path": str(xlsx_path),
        "audio_dir": str(audio_dir),
        "actions": []
    }

    print(f"Committing changes. Writing transaction log to: {transaction_log_path}")
    
    success_count = 0
    fail_count = 0

    for action in proposed_actions:
        if action["status"] not in ("PENDING", "CONFLICT RESOLVED"):
            continue

        src = Path(action["source_path"])
        dst = Path(action["dest_path"])
        artist = action["artist"]
        title = action["title"]

        print(f"Processing Row {action['row']}: {src.name} -> {dst.name}")

        # 1. Update Tags
        tag_ok, tag_err = update_tags(src, artist, title)
        if not tag_ok:
            print(f"  [WARN] Failed to update tags: {tag_err}")
            # We still proceed with renaming even if tag update fails, but we log it

        # 2. Rename File
        try:
            os.rename(src, dst)
            transaction_log["actions"].append({
                "row": action["row"],
                "old_path": str(src),
                "new_path": str(dst),
                "artist": artist,
                "title": title,
                "tag_updated": tag_ok
            })
            success_count += 1
            print("  [OK] Renamed successfully.")
        except Exception as e:
            print(f"  [ERROR] Failed to rename: {e}")
            fail_count += 1

        # Write transaction log incrementally to prevent data loss if interrupted
        try:
            with open(transaction_log_path, 'w', encoding='utf-8') as f:
                json.dump(transaction_log, f, indent=2)
        except Exception as e:
            print(f"  [CRITICAL] Failed to write transaction log: {e}")

    print(f"\nExecution complete. Success: {success_count} | Failed: {fail_count}")
    print(f"Transaction log saved to: {transaction_log_path}")


if __name__ == "__main__":
    main()
